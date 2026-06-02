"""
excel_utils.py

负责 Excel 文件初始化与数据写入。
Excel 文件包括两个 sheet：
1. sensor_data：保存板端上传的原始传感器数据
2. sleep_result：保存 PC 端分类结果
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from protocol_config import (
    EXCEL_FILE,
    SENSOR_SHEET,
    RESULT_SHEET,
    SENSOR_FIELDS,
    RESULT_FIELDS,
    SLEEP_STATE_NAME,
)


def _format_header(ws):
    """
    设置表头样式。
    ws 是一个 worksheet。
    """

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_basic_width(ws, fields):
    """
    设置基础列宽，方便直接打开 Excel 查看。
    """

    for index, field in enumerate(fields, start=1):
        width = max(12, min(24, len(field) + 4))
        ws.column_dimensions[get_column_letter(index)].width = width


def init_excel():
    """
    初始化 Excel 文件。

    如果 sleep_monitor_data.xlsx 不存在：
        创建 Excel
        创建 sensor_data sheet
        创建 sleep_result sheet

    如果文件已经存在：
        直接返回，不覆盖原来的数据
    """

    path = Path(EXCEL_FILE)

    if path.exists():
        return

    wb = Workbook()

    # 第一个 sheet：sensor_data
    ws_sensor = wb.active
    ws_sensor.title = SENSOR_SHEET
    ws_sensor.append(SENSOR_FIELDS)
    _format_header(ws_sensor)
    _set_basic_width(ws_sensor, SENSOR_FIELDS)

    # 第二个 sheet：sleep_result
    ws_result = wb.create_sheet(RESULT_SHEET)
    ws_result.append(RESULT_FIELDS)
    _format_header(ws_result)
    _set_basic_width(ws_result, RESULT_FIELDS)

    wb.save(EXCEL_FILE)


def append_sensor_data(data: dict):
    """
    追加一条 sensor_data 到 Excel。

    data 是从 JSON 解析得到的 Python 字典。
    """

    init_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SENSOR_SHEET]

    row = []

    for field in SENSOR_FIELDS:
        row.append(data.get(field, ""))

    ws.append(row)
    wb.save(EXCEL_FILE)


def append_sleep_result(result: dict):
    """
    追加一条 sleep_result 到 Excel。

    result 是 PC 端分类后的结果字典。
    """

    init_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[RESULT_SHEET]

    code = result.get("sleep_state_code", "")

    row_data = dict(result)
    row_data["sleep_state_name"] = SLEEP_STATE_NAME.get(code, "未知状态")

    row = []

    for field in RESULT_FIELDS:
        row.append(row_data.get(field, ""))

    ws.append(row)
    wb.save(EXCEL_FILE)